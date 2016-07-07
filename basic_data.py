from scheduler.models import Venue, Location, Roster, Challenge, Training, Coach
from con_event.models import Country, State, Con, Registrant, Blackout, Blog,LOCATION_TYPE
from django.contrib.auth.models import Group, User
from swingtime.models import Occurrence,TrainingRoster
from django.core.exceptions import ObjectDoesNotExist
from django.db.models import Q
from django.utils import timezone
import csv
import os
#import string
import datetime
from datetime import datetime
from rcreg_project.settings import BASE_DIR
from rcreg_project.extras import remove_punct,ascii_only,ascii_only_no_punct
import openpyxl
import collections
from random import choice
import random, string

static_path=BASE_DIR+"/static/data/"
import_path=static_path+'unformatted/'
export_path=static_path+'exported/'

data_columns=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X',
    'Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN']


#con=Con.objects.get(year="2016")
#chalo_dict=get_chal_os(con)
def get_chal_os(con):
    """For printing out Excel backup sheets.
    Gets all scheduled Challenge Occurrances for con, returns dict w/ K,V  of date, chal list."""
    chalo_dict={}
    chalos=Occurrence.objects.filter(start_time__gte=con.start,end_time__lte=con.end).exclude(challenge=None)
    for c in chalos:
        if c.start_time.date() not in chalo_dict:
            chalo_dict[c.start_time.date()]=[c]
        else:
            tmp=chalo_dict.get(c.start_time.date())
            tmp.append(c)
            chalo_dict[c.start_time.date()]=list(tmp)

    return chalo_dict

#traino_dict=get_train_os(con)
def get_train_os(con):
    """For printing out Excel backup sheets.
    Gets all scheduled Challenge Occurrances for con, returns dict w/ K,V  of date, chal list."""
    o_dict={}
    tos=Occurrence.objects.filter(start_time__gte=con.start,end_time__lte=con.end).exclude(training=None)
    for c in tos:
        if c.start_time.date() not in o_dict:
            o_dict[c.start_time.date()]=[c]
        else:
            tmp=o_dict.get(c.start_time.date())
            tmp.append(c)
            o_dict[c.start_time.date()]=list(tmp)
    return o_dict

#make_chal_backup(chalo_dict)
def make_chal_backup(acto_dict):
    """Takes in dict of k,b day, occurrence list, to make basic excel sheet backups.
    directory tree: Date, location, chal time+name"""
    BASE_DIR=os.getcwd()
    now=timezone.now()
    nowstr=now.strftime('%m %d %Y')
    com_str="Competition (printed %s)"%(nowstr)
    comp_path=os.path.join(BASE_DIR, com_str)

    if not os.path.isdir(comp_path):
        os.makedirs(comp_path)

    for k,v in acto_dict.iteritems():
        daystr=k.strftime('%m-%d-%Y')
        daypath=os.path.join(comp_path, daystr)
        if not os.path.isdir(daypath):
            os.makedirs(daypath)

        for acto in v:
            oloc=acto.location
            loc_folder="%s (%s)"%(oloc.abbrv,daystr)
            loc_path=os.path.join(daypath, loc_folder)
            #loc_path=os.path.join(daypath, oloc.abbrv)
            if not os.path.isdir(loc_path):
                os.makedirs(loc_path)

            timestr=acto.start_time.strftime('%H %M %p ')
            xlfilename=timestr+(acto.name)+".xlsx"
            fullfilename=os.path.join(loc_path, xlfilename)

            wb = openpyxl.Workbook()
            sheet = wb.active

            sheet["A1"].value = acto.challenge.roster1.name
            sheet["B1"].value = "VS"
            sheet["C1"].value = acto.challenge.roster2.name

            sheet["A2"].value = acto.location.abbrv
            sheet["B2"].value = acto.start_time.strftime('%H %M %p, %m-%d-%Y')

            sheet["D2"].value = "Printed:"
            sheet["E2"].value = nowstr


            sheet["A4"].value = "TEAM"
            sheet["B4"].value = acto.challenge.roster1.name
            sheet["A5"].value = "COLOR"
            sheet["B5"].value = acto.challenge.roster1.color
            sheet["A6"].value = "CAPTAIN"
            sheet["B6"].value = acto.challenge.roster1.captain.name
            sheet["A7"].value = "SKILL"
            sheet["B7"].value = acto.challenge.roster1.skill_display()
            sheet["A8"].value = "GENDER"
            sheet["B8"].value = acto.challenge.roster1.gender_text()

            sheet["E4"].value = "TEAM"
            sheet["E5"].value = "COLOR"
            sheet["F4"].value = acto.challenge.roster2.name
            sheet["F5"].value = acto.challenge.roster2.color
            sheet["E6"].value = "CAPTAIN"
            sheet["F6"].value = acto.challenge.roster2.captain.name
            sheet["E7"].value = "SKILL"
            sheet["F7"].value = acto.challenge.roster2.skill_display()
            sheet["E8"].value = "GENDER"
            sheet["F8"].value = acto.challenge.roster2.gender_text()

            sheet["A10"].value = "# of players"
            sheet["E10"].value = "# of players"
            sheet["B10"].value = "Skater #"
            sheet["F10"].value = "Skater #"
            sheet["C10"].value = "Skater Name"
            sheet["G10"].value = "Skater Name"

            starti=11
            rno=int(1)
            r1=list(acto.challenge.roster1.participants.all())
            r1.sort(key=lambda x: x.sk8number)
            for r in r1:
                if r==acto.challenge.roster1.captain:
                    if r.sk8name:
                        name=r.sk8name+" (Captain)"
                    else:
                        name="(Captain)"
                else:
                    name=r.sk8name
                sheet["A"+str(starti)].value = str(rno)+"."
                sheet["B"+str(starti)].value = r.sk8number
                sheet["C"+str(starti)].value = name
                rno+=1
                starti+=1

            starti=11
            rno=int(1)
            r2=list(acto.challenge.roster2.participants.all())
            r2.sort(key=lambda x: x.sk8number)
            for r in r2:
                if r==acto.challenge.roster2.captain:
                    if r.sk8name:
                        name=r.sk8name+" (Captain)"
                    else:
                        name="(Captain)"
                else:
                    name=r.sk8name
                sheet["E"+str(starti)].value = str(rno)+"."
                sheet["F"+str(starti)].value = r.sk8number
                sheet["G"+str(starti)].value = name
                rno+=1
                starti+=1

            wb.save(filename = fullfilename)
    print "done making xl files"


#make_train_backup(traino_dict)
def make_train_backup(acto_dict):
    """Takes in dict of k,b day, occurrence list, to make basic excel sheet backups.
    directory tree: Date, location, chal time+name"""
    BASE_DIR=os.getcwd()
    now=timezone.now()
    nowstr=now.strftime('%m %d %Y')
    train_str="Training (printed %s)"%(nowstr)
    comp_path=os.path.join(BASE_DIR, train_str)

    if not os.path.isdir(comp_path):
        os.makedirs(comp_path)

    for k,v in acto_dict.iteritems():
        daystr=k.strftime('%m-%d-%Y')
        daypath=os.path.join(comp_path, daystr)
        if not os.path.isdir(daypath):
            os.makedirs(daypath)

        for acto in v:
            oloc=acto.location
            loc_folder="%s (%s)"%(oloc.abbrv,daystr)
            loc_path=os.path.join(daypath, loc_folder)
            if not os.path.isdir(loc_path):
                os.makedirs(loc_path)

            timestr=acto.start_time.strftime('%H %M %p ')
            xlfilename=timestr+(acto.name)+".xlsx"
            fullfilename=os.path.join(loc_path, xlfilename)

            wb = openpyxl.Workbook()
            sheet = wb.active

            sheet["E3"].value = "Printed: %s"%(nowstr)

            sheet["A1"].value = acto.training.name
            sheet["A2"].value = acto.training.display_coach_names()
            sheet["A3"].value = acto.start_time.strftime('%H %M %p, %m-%d-%Y')
            sheet["A4"].value = acto.location.name

            sheet["B6"].value = "Registration Roster"
            sheet["A7"].value = "Skill:"
            sheet["B7"].value = acto.training.skill_display()
            sheet["A8"].value = "Pass:"
            sheet["B8"].value = acto.training.passes_str()
            sheet["B10"].value = "Name"

            sheet["F6"].value = "Auditing Roster"
            sheet["E7"].value = "Skill:"
            sheet["F7"].value = "ABCD"
            sheet["E8"].value = "Pass:"
            sheet["F8"].value = "MVP, Skater, Offskate" #right? confirm.
            sheet["F10"].value = "Name"

            starti=11
            rno=int(1)
            for r in range(1,61):
                sheet["A"+str(starti)].value = str(rno)+"."
                rno+=1
                starti+=1

            starti=11
            rno=int(1)
            for r in range(1,11):
                sheet["E"+str(starti)].value = str(rno)+"."
                rno+=1
                starti+=1


            wb.save(filename = fullfilename)
    print "done making xl files"





# reg,aud,reg_and_aud,chal,both,neither,rall=rostercheck()

#con=Con.objects.get(year="2016")
#captains=get_captains(con)
def get_captains(con):
    challenges=list(Challenge.objects.filter(con=con, RCaccepted=True))
    captains=[]
    for c in challenges:
        for r in [c.roster1,c.roster2]:
            if r and r.captain and r.captain not in captains:
                captains.append(r.captain)
    return captains

#all_bouts=make_random_bo(con,captains)
def make_random_bo(con,captains):
    date_range=con.get_date_range()
    ampm=["AM","PM"]
    all_bouts=[]

    for c in captains:
        bouts=[]
        while len(bouts)<5:
            ap=choice(ampm)
            day=choice(date_range)
            b,created=Blackout.objects.get_or_create(registrant=c,date=day,ampm=ap)
            bouts.append(b)
            all_bouts.append(b)
            b.save()
    return all_bouts




def randomword(length):
   return ''.join(random.choice(string.lowercase) for i in range(length))


def make_random_acts():
    con=Con.objects.get(year="2016")
    all_reg=list(Registrant.objects.filter(con=con))
    print "len all_reg",len(all_reg)
    coaches=[]
    while len(coaches)<20:
        coaches.append(choice(all_reg))
    #print "coaches",coaches
    print "len coaches",len(coaches)
    captains=coaches[:5]
    while len(captains)<100:
        captains.append(choice(all_reg))
    #print "captains",captains
    print "len captains",len(captains)

    # trainings=[]
    # while len(trainings)<40:
    #     c=choice(coaches)
    #     print c
    #     coach, created=Coach.objects.get_or_create(user=c.user)
    #     print "coach",coach
    #     location_type=choice(LOCATION_TYPE)[0]
    #     t=Training(name=randomword(10),con=con,RCaccepted=True, duration="2",interest=3,location_type=location_type)
    #     trainings.append(t)
    #     t.save()
    #     print t," saved!"
    #     t.coach.add(coach)
    #     t.save()

    challenges=[]
    while len(challenges)<100:
        location_type=choice(LOCATION_TYPE[:3])[0]
        r1=Roster(name=randomword(10),con=con,captain=choice(all_reg))
        r2=Roster(name=randomword(10),con=con,captain=choice(all_reg))
        for r in [r1,r2]:
            r.save()
            while r.participants.count()<20:
                r.participants.add(choice(all_reg))
            r.save()
        c=Challenge(con=con,RCaccepted=True, submitted_on=timezone.now(), interest=3,location_type=location_type,roster1=r1,roster2=r2,captain1accepted=True,captain2accepted=True)
        challenges.append(c)
        c.save()

    #print "challenges",challenges


#con=Con.objects.get(year="2016")
#t2016,c2016,candt =get_year_os(con)
def get_year_os(con):
    print"test to make sure all occurrences are being counted"
    allo=list(Occurrence.objects.all())
    print"all occurrences: ",len(allo)
    o2016=Occurrence.objects.filter(start_time__gte=con.start,end_time__lte=con.end)
    print"2016 con occurrences: ",len(o2016)
    candt=[]
    c2016=[]
    t2016=[]
    for o in o2016:
        if o.training:
            t2016.append(o)
        if o.challenge:
            c2016.append(o)
        if o.training and o.challenge:
            candt.append(o)
    print len(t2016)," training occurrences"
    print len(c2016)," challenge occurrences"
    print len(candt)," training and challenge occurrences, this should be 0"

    return t2016,c2016,candt

#coach_regs=get_coach_reg(con)
def get_coach_reg(con):
    coach_u=[]

    for t in Training.objects.filter(con=con,RCaccepted=True):
        for c in t.coach.all():
            coach_u.append(c.user)

    coach_regs=list(Registrant.objects.filter(user__in=coach_u, con=con))
    print "%s coaches for %s"%(str(len(coach_regs)), str(con))

    return coach_regs


def check_chal_conflicts(con):
    chal2016=Occurrence.objects.filter(start_time__gte=con.start,end_time__lte=con.end).exclude(challenge=None)
    rosters=[]
    rostered_sk8ers=[]

    conflict=[]
    free=[]

    for c in chal2016:
        for r in [c.roster1,c.roster2]:
            if r and r not in rosters:
                rosters.append(r)
    print "%s scheduled rosters for %s"%(str(len(rosters)), str(con))
    for r in rosters:
        for s in r.participants.all():
            if s not in rostered_sk8ers:
                rostered_sk8ers.append(s)
    print "%s scheduled skaters for %s"%(str(len(rostered_sk8ers)), str(con))
    #######here's where i'd do sk8er.get_occurrences, but haven't tested it yet.





#sk8er=Registrant.objects.get(pk=72)
#conflict,free =check_sk8er_schedule_conflict(sk8er)
def check_sk8er_schedule_conflict(sk8er):
    conflict=[]
    free=[]
    print "\n\nchecking schedule for ",sk8er, "pk ",sk8er.pk

    sk8er_ros=sk8er.roster_set.all()
    sk8er_coach_list=list(Coach.objects.filter(user=sk8er.user))
    if len(sk8er_coach_list)>0:
        #print "found ",len(sk8er_coach_list)," coach"
        sk8er_coach=sk8er_coach_list[0]
        sk8er_train=sk8er_coach.training_set.filter(con=sk8er.con)
    else:
        sk8er_train=[]

    sk8er_chal=list(Challenge.objects.filter(RCaccepted=True).filter(Q(roster1__in=sk8er_ros)|Q(roster2__in=sk8er_ros)))

    sk8ero=list(Occurrence.objects.filter( Q(challenge__in=sk8er_chal)|Q(training__in=sk8er_train)))
    print "%s is in %s scheduled occurrences" %(sk8er,str(len(sk8ero)) )

    for o in sk8ero:

        concurrent=Occurrence.objects.filter(start_time__lt=(o.end_time + datetime.timedelta(minutes=30)),end_time__gt=(o.start_time - datetime.timedelta(minutes=30))).filter(Q(challenge__in=sk8er_chal)|Q(training__in=sk8er_train)).exclude(pk=o.pk)

        if len(concurrent)<1:
            free.append(o)
        else:
            conflict.append(o)

    print "%s conflicting and %s free scheduled occurrences for %s" %( str(len(conflict)) , str(len(free)), str(sk8er.name))
    print"Conflicts: "
    for c in conflict:
        print "%s %s - %s"%(str(c.name), c.start_time.strftime("%a %B %d %I:%-M %p"), c.end_time.strftime("%I:%-M %p") )
    return conflict,free


def roster_skills_check():
    con=Con.objects.get(year="2016")
    mismatch=[]
    approved=Challenge.objects.filter(RCaccepted=True, con=con)
    for c in approved:
        r1=c.roster1.opponent_skills_allowed()
        r2=c.roster2.opponent_skills_allowed()
        if (c.roster1.skill_display() not in r2) or  (c.roster2.skill_display() not in r1):
            mismatch.append(c)

    print "all approved: ",len(approved)
    print "mismatches: ",len(mismatch)
    for c in mismatch:
        print "\n"
        print "MISMATCH: ",c
        r1skills=[]
        r2skills=[]
        print c.roster1.pk, c.roster1, c.roster1.skill_display()
        for s in c.roster1.participants.all():
            print s, s.skill
            if s.skill not in r1skills:
                r1skills.append(s.skill)
        print "skills list: ",r1skills
        print "\n"
        print c.roster2.pk, c.roster2, c.roster2.skill_display()
        for s in c.roster2.participants.all():
            print s, s.skill
            if s.skill not in r2skills:
                r2skills.append(s.skill)
        print "skills list: ",r2skills
        print "\n"


def rostercheck():
    reg=[]
    aud=[]
    reg_and_aud=[]
    chal=[]

    both=[]
    neither=[]
    rall=[]

    for r in Roster.objects.all():
        cs=list(r.roster1.all())+list(r.roster2.all())
        if r.registered and r.auditing and (len(cs)>=1):
            rall.append(r)
        elif not r.registered and not r.auditing and (len(cs)<1):
            neither.append(r)
        elif (r.registered or r.auditing) and (len(cs)>=1):
            both.append(r)
        elif not (r.registered or r.auditing) and (len(cs)>=1):
            chal.append(r)
        elif r.registered and r.auditing:
            reg_and_aud.append(r)
        elif not r.registered and r.auditing:
            reg.append(r)
        elif r.registered and not r.auditing:
            aud.append(r)

    print "reg: ",len(reg)
    print "aud: ",len(aud)
    print "reg_and_aud: ",len(reg_and_aud)
    print "chal: ",len(chal)
    print "both: ",len(both)
    print "neither: ",len(neither)
    print "rall: ",len(rall)

    return reg,aud,reg_and_aud,chal,both,neither,rall


def get_gametypes():
    con=Con.objects.get(year="2016")
    gametypedict={}

    for c in Challenge.objects.filter(con=con, RCaccepted=True):
        if c.gametype not in gametypedict:
            gametypedict[c.gametype]=[c]
        else:
            templist=gametypedict.get(c.gametype)
            templist.append(c)
            gametypedict[c.gametype]=list(templist)

    for k,v in gametypedict.iteritems():
        print k
        for item in v:
            print item.pk, item.name, item.gametype, item.duration, item.is_a_game
    print "\nSUMMARY\n"
    for k,v in gametypedict.iteritems():
        print k, ": ",len(v)


def get_duration():
    con=Con.objects.get(year="2016")
    durdict={}

    for c in Challenge.objects.filter(con=con, RCaccepted=True):
        if c.duration not in durdict:
            durdict[c.duration]=[c]
        else:
            templist=durdict.get(c.duration)
            templist.append(c)
            durdict[c.duration]=list(templist)
    print "\nChallenges\n"
    for k,v in durdict.iteritems():
        print k
        for item in v:
            print item.pk, item.name, item.gametype, item.duration, item.is_a_game
    print "\nChallenge SUMMARY\n"
    for k,v in durdict.iteritems():
        print k, ": ",len(v)

    durdictTR={}
    for c in Training.objects.filter(con=con, RCaccepted=True):
        if c.duration not in durdictTR:
            durdictTR[c.duration]=[c]
        else:
            templist=durdictTR.get(c.duration)
            templist.append(c)
            durdictTR[c.duration]=list(templist)
    print "\nTrainings\n"
    for k,v in durdictTR.iteritems():
        print k
        for item in v:
            print item.pk, item.name, item.duration
    print "\nTraining SUMMARY\n"
    for k,v in durdictTR.iteritems():
        print k, ": ",len(v)



def get_idosyncracies():
    perfect=[]
    imperfect=[]
    con=Con.objects.get(year="2016")

    for c in Challenge.objects.filter(con=con, RCaccepted=True):
        if c.gametype!="6GAME" or c.is_a_game:
            print "%s is_a_game or 6GAME, and duration is %s"%(c.name, c.duration)
            if float(c.duration)<1.0:
                print  "%s duration is %s"%(c.name, c.duration)
                if c not in imperfect:
                    imperfect.append(c)

            if c.is_a_game and c.gametype!="6GAME":
                print "%s is_a_game, but is not 6GAME"%(c.name)
                if c not in imperfect:
                    imperfect.append(c)
            if c.gametype=="6GAME" and not c.is_a_game:
                print "%s is 6GAME, but not is_a_game"%(c.name)
                if c not in imperfect:
                    imperfect.append(c)

        if c.gametype=="6GAME" and c.is_a_game and float(c.duration)>=1.0:
            if c not in perfect:
                perfect.append(c)
        elif c.gametype!="6GAME" and not c.is_a_game and float(c.duration)<=1.0:
            if c not in perfect:
                perfect.append(c)

        if c.gametype!="6GAME" and not c.is_a_game and float(c.duration)>=1.0:
            print "%s is not a game, but duration is %s"%(c.name, c.duration)
            if c not in imperfect:
                imperfect.append(c)

    print "Perfect: ",len(perfect)
    print "Imperfect: ",len(imperfect)
    return perfect, imperfect


def test_interest_default():
    con=Con.objects.get(year="2016")
    for t in Training.objects.filter(con=con, RCaccepted=True):
        if not t.interest:
            print t.name, "Coaches: ",t.display_coach_names(),t.get_default_interest()#don'tsave!
        else:
            print "%s has an interest, it's %s"%(t.name,str(t.interest))
    print " "
    for c in Challenge.objects.filter(con=con, RCaccepted=True):
        if not c.interest:
            print c.name, c.get_default_interest()#don'tsave!
        else:
            print "%s has an interest, it's %s"%(t.name,str(t.interest))

#con=Con.objects.get(year="2016")
#qset=list(Training.objects.filter(con=con))
#loc=show_loc(qset)
def show_loc(qset):
    loc={}
    for t in qset:
        if t.location_type not in loc:
            loc[t.location_type]=[t]
        else:
            temp=loc.get(t.location_type)
            temp.append(t)
            loc[t.location_type]=list(temp)
    return loc


def wb_or_str(xlfile):
    """cehcks is xlfile input is a strong of name, or wb object. either makes or returns object."""
    if isinstance(xlfile , basestring):
        #if a string of file name is entered
        wb = openpyxl.load_workbook(xlfile)
    else:
        # if a wb object is entered
        wb = xlfile
    return wb

#xlfile=(static_path+'myheader.xlsx')
#xlfile=(static_path+'BPTheader.xlsx')
def get_header(xlfile):
    wb=wb_or_str(xlfile)
    sheet = wb.get_active_sheet()
    #sheet=wb.get_sheet_by_name('downloadreports-1')#this only works for RollerTron.xlsx
    header=collections.OrderedDict()
    for row in range(1, sheet.get_highest_row() + 1):
        for  c in data_columns:#global vairable, look at top of file
            location=c+str(row)
            data=sheet[location].value
            header[c]=data
    return header

def write_wb(target_location,target_name,od_list,header):
    """od_list is a list of ordered dicts, w/ k being target column in an excel sheet (A, B, etc) and v being the value,
    Each od represents 1 row."""

    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    r=int(1)
    for od in od_list:
        while r<2:#write the header
            for k,v in header.items():
                location=str(k)+str(r)
                sheet[location].value = v
            r+=int(1)

        for k,v in od.items():
            location=str(k)+str(r)
            sheet[location].value = v
        r+=int(1)
    wb.save(target_location+target_name)
    return wb

def make_excel_odict_list(xlfile):
    """Takes in excel file of BPT registrant data, turns each row into an ordered dict, returns list of all ordered dicts"""
    wb=wb_or_str(xlfile)
    sheet = wb.get_active_sheet()
    #sheet=wb.get_sheet_by_name('downloadreports-1')#this only works for RollerTron.xlsx
    all_data=[]
    highest_row=sheet.get_highest_row()
    for row in range(2, sheet.get_highest_row() + 1):
        data_holding_dict = collections.OrderedDict()
        for  c in data_columns:
            location=c+str(row)
            data=sheet[location].value
            data_holding_dict[c]=data
        all_data.append(data_holding_dict)
    #so by now all_data has a shit ton of stuff
    return all_data

#xlfile=(export_path+'SingleEmailRegistrants.xlsx')
def find_incompletes(xlfile):
    """this assumes BPT header, not my header"""
    all_data=make_excel_odict_list(xlfile)
    no_sk8name=[]
    no_real_name=[]
    complete_entries=[]

    for od in all_data:
        sk8name=od.get("AC")
        first_name=od.get("AA")
        last_name=od.get("Z")

        if not sk8name:
            no_sk8name.append(od)
        if not first_name or not last_name:
            no_real_name.append(od)
        if sk8name and first_name and last_name:
            complete_entries.append(od)

    header=get_header((static_path+'BPTheader.xlsx'))
    date_str=datetime.date.today().strftime("%B %d %Y")#eg 'July 23 2010'

    if len(no_sk8name)>0:
        name_str='NoSk8NameReg '+ date_str +'.xlsx'
        no_sk8name_file=write_wb(export_path,name_str,no_sk8name,header)
    else:
        no_sk8name_file=None

    if len(no_real_name)>0:
        name_str='NoRealNameReg '+ date_str +'.xlsx'
        no_real_name_file=write_wb(export_path,name_str,no_real_name,header)
    else:
        no_real_name_file=None

    if len(complete_entries)>0:
        name_str='CompleteReg '+ date_str +'.xlsx'
        complete_entries_file=write_wb(export_path,name_str,complete_entries,header)
    else:
        complete_entries_file=None

    return no_sk8name_file,no_real_name_file,complete_entries_file


#xlfile=(import_path+"RollerTron.xlsx")
def email_dupes(xlfile):
    """Takes in list of ordered dicts from BPT Excel sheet, shits out 2 excels: 1 of people who entered unique emails,
    1 of peolpe who are attached to an email that is used more than once."""
    all_data=make_excel_odict_list(xlfile)
    email_list=[]
    last_email=None
    good_emails=[]
    bad_emails=[]
    long_emails=[]

    for od in all_data:
        email2=od.get("AB")
        email_list.append(email2)
        try:
            if len(email2)>=30:
                print "long email ",email2
                long_emails.append(od)
        except:
            print "error with ",od

    for od in all_data:
        email2=od.get("AB")
        if int(email_list.count(email2))>1:
            bad_emails.append(od)
        else:
            good_emails.append(od)

    header=get_header((static_path+'BPTheader.xlsx'))
    date_str=datetime.date.today().strftime("%B %d %Y")#eg 'July 23 2010'
    if len(good_emails)>0:
        name_str='SingleEmailRegistrants '+ date_str +'.xlsx'
        single_file=write_wb(export_path,name_str,good_emails,header)
    else:
        single_file=None

    if len(bad_emails)>0:
        name_str='EmailDupeRegistrants '+ date_str +'.xlsx'
        dupe_file=write_wb(export_path,name_str,bad_emails,header)
    else:
        dupe_file=None

    if len(long_emails)>0:
        name_str='LONGEmailRegistrants '+ date_str +'.xlsx'
        dupe_file=write_wb(export_path,name_str,long_emails,header)

    return single_file,dupe_file

def sort_BPT_excel(target_file):
    """aggregates the cleaner funcitons, so i can enter the big BPT excel and shit out: good/bad emails, 2 incomplete name files, 1 complete name file"""
    BPT_header = get_header((static_path+'BPTheader.xlsx'))
    single_file,dupe_file=email_dupes(target_file)
    no_sk8name_file,no_real_name_file,complete_entries_file=find_incompletes(single_file)

    return single_file, dupe_file, no_sk8name_file, no_real_name_file, complete_entries_file

#from basic_data import*
#target_file=(import_path+'RollerTron Attendee 050416.xlsx')
#con=Con.objects.get(year="2016")
#single_file, dupe_file, no_sk8name_file, no_real_name_file, complete_entries_file=sort_BPT_excel(target_file)
def import_from_excel(complete_entries_file,con):
    """This assumes that I've already checked for duplicate emails and lack of name, sk8name.
    This is data that could be ready for import via Django import/export, but I think this will be faster.
    This is coming from BPT format, not mine"""
    all_data=make_excel_odict_list(complete_entries_file)
    error_list=[]
    success_list=[]
    repeat_email_list=[]
    for od in all_data:
        skill=od.get("AF")
        gender=od.get("AG")
        pass_type=od.get("V")
        email=od.get("AB")
        if not email:
            email=od.get("Q")
        first_name =od.get("AA")
        if not first_name:
            first_name =od.get("I")
        last_name = od.get("Z")
        if not last_name:
            last_name = od.get("H")
        sk8name = od.get("AC")
        sk8number =str(od.get("AE"))
        country_name = od.get("O")
        state_name=od.get("M")
        BPT_Ticket_ID=od.get("A")

        affiliation=od.get("AH")
        ins_carrier=od.get("AI")
        ins_number=od.get("AJ")
        age_group=od.get("AK")
        favorite_part=od.get("AM")
        volunteer=od.get("AL")
        for att in [first_name,last_name,sk8name,sk8number,affiliation,ins_carrier,ins_number,favorite_part,volunteer]:
            cleaned_att=ascii_only_no_punct(att)
            str_att=str(cleaned_att)
            if len(str_att)>100:
                att=str_att[:100]

        if country_name:
            try:
                country=Country.objects.get(name__iexact=country_name)
            except:
                country=None
        else:
            country=None

        if state_name:
            try:
                state=State.objects.get(slugname=state_name)
            except:
                state=None
        else:
            state=None

        try:#this is the overarching try, any failure will send to error_list
            this_reg=None
            try:#stating with most matches, broadening, until i'm sure it doesn't exist
                this_reg=Registrant.objects.get(con=con, email=email,first_name=first_name,last_name=last_name,sk8name=sk8name)
                print "found %s, first try"%(this_reg)
            except ObjectDoesNotExist:
                #first try: con and email match, and EITHER f/l name or ska8name
                reg_q=Registrant.objects.filter(con=con, email=email).filter(Q(first_name__iexact=first_name,last_name__iexact=last_name)|Q(sk8name__iexact=sk8name))
                if reg_q.count()==1:
                    this_reg=reg_q[0]
                    print "found %s, second try"%(this_reg)
                else:#allow for same person, different email
                    reg_q=Registrant.objects.filter(con=con).filter(Q(first_name__iexact=first_name,last_name__iexact=last_name)|Q(sk8name__iexact=sk8name))
                    if reg_q.count()==1:
                        this_reg=reg_q[0]
                        print "found %s, third try w/ diff email"%(this_reg)
                    else:
                        try:
                            Registrant.objects.get(con=con, email=email)
                            print "email exists"
                            repeat_email_list.append(od)
                        except ObjectDoesNotExist:
                            #here's where I think doesn't exist, make a new One . if repeat email, will fail upon save
                            print "think doesn't exist"
                            this_reg=Registrant(con=con, email=email,first_name=first_name,last_name=last_name)

            if this_reg:#ie if no repeat email
                attr_dict={'sk8name':sk8name,'sk8number':sk8number,'skill':skill,"gender":gender,'pass_type':pass_type,'first_name':first_name,'last_name':last_name,
                    'country':country,'state':state,'BPT_Ticket_ID':BPT_Ticket_ID,'affiliation':affiliation,'ins_carrier':ins_carrier,'ins_number':ins_number,'age_group':age_group,
                    'favorite_part':favorite_part,'volunteer':volunteer}
                for k,v in attr_dict.iteritems():
                    print k," is ",v
                    #value = getattr(this_reg, k)
                    #if v and not value:
                    if v:
                        print "setting or updating",this_reg,"s ",k
                        setattr(this_reg, k, v)
                    elif not v:
                        print "od doesn't have ",k
                    # elif value:
                    #     print this_reg,"already has a ",k,": ",value
                        #NEW NOTE this was making everyone have default MVP Female setting


                this_reg.save()
                this_reg.save()#think I have to do twice tomake user? I forgot.
                success_list.append(od)
                print this_reg," Succesfully made"
        except:
            print "Fail: "
            error_list.append(od)
            print this_reg


    header=get_header((static_path+'BPTheader.xlsx'))
    date_str=datetime.date.today().strftime("%B %d %Y")#eg 'July 23 2010'

    if success_list:
        name_str='RegistrantsMade '+ date_str +'.xlsx'
        write_wb(export_path,name_str,success_list,header)
        print "success list written"
    if error_list:
        name_str='RegistrantFAIL '+ date_str +'.xlsx'
        write_wb(export_path,name_str,error_list,header)
        print "error list written"
    if repeat_email_list:
        name_str='RegistrantREPEATEMAILFAIL '+ date_str +'.xlsx'
        write_wb(export_path,name_str,repeat_email_list,header)
        print "repeat_email_list written"
