from scheduler.models import Venue, Location, Roster, Challenge, Training, Coach
from con_event.models import Country, State, Con, Registrant, Blog
from django.contrib.auth.models import Group, User
import csv
import os
import datetime
from rcreg_project.settings import BASE_DIR
from rcreg_project.extras import remove_punct,ascii_only,ascii_only_no_punct
import openpyxl
import collections

static_path=BASE_DIR+"/static/"
old_rc_path=static_path+'RC2015/unformatted/'
new_rc_path=static_path+'RC2015/exported/'

def find_retards():
#this works but it lets the first of a repeated email get away
    #csvfile=(old_rc_path+"RollerTron.csv")
    xlfile=(old_rc_path+"RollerTron.xlsx")
    wb = openpyxl.load_workbook(xlfile)
    #sheet = wb.get_active_sheet()
    sheet=wb.get_sheet_by_name('downloadreports-1')
    all_data=[]
    highest_row=sheet.get_highest_row()
    #print "highest_row",highest_row
    for row in range(2, sheet.get_highest_row() + 1):
        #print "row: ",row
        data_columns=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X',
            'Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN']
        data_holding_dict = collections.OrderedDict()
        for  c in data_columns:
            location=c+str(row)
            #print "location",location
            data=sheet[location].value
            #print "data: ",data
            data_holding_dict[c]=data
            #print "data_holding_dict",data_holding_dict
        all_data.append(data_holding_dict)
    #so by now all_data has a shit ton of stuff


    email_list=[]
    last_email=None
    good_emails=[]
    bad_emails=[]
    for od in all_data:
        email2=od.get("AB")
        print "email2",email2
        if (email2 in email_list) or email2==last_email:
            bad_emails.append(od)
            print "bad email",od
        else:
            print "good email",od
            good_emails.append(od)
            email_list.append(email2)
            last_email=email2
    print "good emails:", len(good_emails)
    print "bad emails:", len(bad_emails)
    print "total: ",len(good_emails)+len(bad_emails)

    goodwb = openpyxl.Workbook()
    sheet = goodwb.get_active_sheet()
    r=int(0)
    print "GOOD EMAILS"
    for od in good_emails:
        r+=int(1)
        for k,v in od.items():
            location=str(k)+str(r)
            print "location",location
            print "data",v
            #sheet[location]= v
            sheet[location].value = v
    goodwb.save(old_rc_path+'GoodRegistrants.xlsx')

    badwb = openpyxl.Workbook()
    sheet = badwb.get_active_sheet()
    r=int(0)
    print "BAD EMAILS"
    for od in bad_emails:
        r+=int(1)
        for k,v in od.items():
            location=str(k)+str(r)
            print "location",location
            print "data",v
            #sheet[location] = v
            sheet[location].value = v
    badwb.save(old_rc_path+'BadRegistrants.xlsx')
