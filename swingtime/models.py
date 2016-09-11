import datetime
from dateutil import rrule
from django.contrib.auth.models import User,Group
from django.utils import timezone
from django.utils.translation import ugettext_lazy as _
from django.utils.encoding import python_2_unicode_compatible
from django.utils.dateparse import parse_datetime,parse_time
from django.contrib.contenttypes.models import ContentType
from django.db import models
from django.db.models import Q,F
from django.conf import settings
from django.core.urlresolvers import reverse
from django.core.exceptions import ValidationError, NON_FIELD_ERRORS

from scheduler.models import Location, Challenge, Training,INTEREST_RATING
from scheduler.app_settings import DEFAULT_REG_CAP, DEFAULT_AUD_CAP

from con_event.models import Blackout,Registrant,SKILL_LEVEL_TNG
from rcreg_project.settings import BIG_BOSS_GROUP_NAME,LOWER_BOSS_GROUP_NAME
from rcreg_project.extras import ascii_only_no_punct

from random import choice
import openpyxl
from openpyxl.styles import PatternFill
try:
    from django.contrib.contenttypes.fields import GenericForeignKey, GenericRelation
except ImportError:
    from django.contrib.contenttypes.generic import GenericForeignKey, GenericRelation

from swingtime.conf import settings as swingtime_settings


#####NOTE to self (Dahmer):
#temporarily got rid of Note, description field.
#am thinking of getitng rid of title and Event type.
#title is probably okay as long as I keep funciton to return chal or training
#only get rid of event type if you know you'll never want to use the css helper function.
#either way, need to resolve either of these in utils if get rid of them/

#===============================================================================
@python_2_unicode_compatible
class EventType(models.Model):
    '''
    Simple ``Event`` classifcation.
    '''
    abbr = models.CharField(_('abbreviation'), max_length=4, unique=True)
    label = models.CharField(_('label'), max_length=50)

    #===========================================================================
    class Meta:
        verbose_name = _('event type')
        verbose_name_plural = _('event types')

    #---------------------------------------------------------------------------
    def __str__(self):
        return self.label


#===============================================================================
@python_2_unicode_compatible
class Event(models.Model):
    '''
    Container model for general metadata and associated ``Occurrence`` entries.
    '''
    #title = models.CharField(_('title'), max_length=32,null=True,blank=True)
    #event_type = models.ForeignKey(EventType, verbose_name=_('event type'),null=True,blank=True)

    training=models.ForeignKey(Training,null=True,blank=True,on_delete=models.SET_NULL)
    challenge=models.ForeignKey(Challenge,null=True,blank=True,on_delete=models.SET_NULL)

    #===========================================================================
    class Meta:
        verbose_name = _('event')
        verbose_name_plural = _('events')
        #ordering = ('title', )

    #---------------------------------------------------------------------------
    def get_activity(self):
        '''
        Dahmer custom. Returns training or activity related to Event.
        '''
        if self.training:
            activity=self.training
        elif self.challenge:
            activity=self.challenge
        else:
            activity=None
        return activity

    #---------------------------------------------------------------------------
    def __str__(self):
        activity=self.get_activity()
        if activity:
            return activity.name
        else:
            return "no challenge or training yet"


#===============================================================================
class OccurrenceManager(models.Manager):

    use_for_related_fields = True

    #---------------------------------------------------------------------------
    def daily_occurrences(self, dt=None):
        '''
        Returns a queryset of for instances that have any overlap with a
        particular day.

        * ``dt`` may be either a datetime.datetime, datetime.date object, or
          ``None``. If ``None``, default to the current day.

        '''
        dt = dt or datetime.datetime.now()
        start = datetime.datetime(dt.year, dt.month, dt.day)
        end = start.replace(hour=23, minute=59, second=59)
        qs = self.filter(
            models.Q(
                start_time__gte=start,
                start_time__lte=end,
            ) |
            models.Q(
                end_time__gte=start,
                end_time__lte=end,
            ) |
            models.Q(
                start_time__lt=start,
                end_time__gt=end
            )
        )

        return qs

    #---------------------------------------------------------------------------
    def gather_possibles(self, con,all_act_data):
        """Meant to make Otto run faster by moving all db queries to 1 place.
        Takes in activity list, either list of approved but unscheduled challenges or trainings
        gets all possible Occurrences for entire group, divides up by which would suit each activity
        doesbn't sort by conflict yet"""
        #0 db hits?!? wtf?
        #print "gather possibles experiment"
        #print "dbc1:", len(dbconnection.queries)
        venues=con.venue.all()
        all_locations=Location.objects.filter(venue__in=venues) #this works. should i select/prefetch anything?
        #date_range=con.get_date_range()

        #base_q=Occurrence.objects.filter(challenge=None,training=None,start_time__gte=con.start, end_time__lte=con.end)
        base_q=list(Occurrence.objects.filter(challenge=None,training=None,start_time__gte=con.start, end_time__lte=con.end).select_related('location'))

        #print "dbc2:", len(dbconnection.queries)

        possibles=all_act_data.keys()

        for act in possibles:
            this_act_data={}

            ############ start gathering locaiton per activity###############
            ########this is where it's beng evaluated, i bet. If I oculd sort ahead of time I'd massively cut down on db trips.
            if act.location_type =='Flat Track':
                if act.is_a_training():#if this is a training
                    act_locations=all_locations.filter(venue__in=venues, location_type='Flat Track', location_category="Training")

                elif act.is_a_challenge():
                    if act.is_a_game or float(act.duration)>=1:#has to be in C1
                        act_locations=all_locations.filter(venue__in=venues, location_type='Flat Track', location_category="Competition Any Length")
                    else:#can be n C1 or C2
                        act_locations=all_locations.filter(venue__in=venues, location_type='Flat Track', location_category__in=["Competition Half Length Only","Competition Any Length"])

            elif act.location_type == 'EITHER Flat or Banked Track':
                if act.is_a_training():#if this is a training
                    act_locations=all_locations.filter(location_category__in=["Training","Training or Competition"], venue__in=venues, location_type__in=['Flat Track','Banked Track'])
                elif act.is_a_challenge():
                    act_locations=all_locations.filter(location_category__in=["Training or Competition","Competition Half Length Only","Competition Any Length"],venue__in=venues, location_type__in=['Flat Track','Banked Track'])
            else:
                act_locations=all_locations.filter(venue__in=venues, location_type=act.location_type)

            this_act_data["locations"]=act_locations
            ############ end gathering locaiton per activity###############

            ############ start interest, activity type, per activity###############
            if act.interest:
                proxy_interest=act.interest
            else:
                proxy_interest=act.get_default_interest()

            if act.is_a_training():#if this is a training
                proxy_interest=abs(6-proxy_interest)#to make high demand classes in low interest timeslots and vice versa
            elif act.is_a_challenge():
                this_act_data["proxy_interest"]=proxy_interest
            duration=float(act.duration)
            dur_delta=int(duration*60)
            this_act_data["dur_delta"]=dur_delta
            ############ end interest, activity type, per activity###############
            #print "dbc2.5:", len(dbconnection.queries)

            #act_os=base_q.filter(interest__in=[proxy_interest-1,proxy_interest,proxy_interest+1], location__in=act_locations,end_time=F('start_time') + timedelta(minutes=dur_delta))
            act_os=[]
            for o in base_q:
                if (o.interest in [proxy_interest-1,proxy_interest,proxy_interest+1]) and (o.location in act_locations) and (o.end_time==(o.start_time + datetime.timedelta(minutes=dur_delta)) ) :
                    act_os.append(o)


            this_act_data["act_os"]=act_os
            #print "dbc2.6:", len(dbconnection.queries)
            interestexact=[]
            interestremoved=[]

            for o in act_os:#this is where it gets evaluated, 1 for every acot
                if o.interest==proxy_interest:
                    interestexact.append(o)
                else:
                    interestremoved.append(o)
            this_act_data["interestexact"]=interestexact
            this_act_data["interestremoved"]=interestremoved
            #print "dbc2.7:", len(dbconnection.queries)
            #update dict w/ new data
            old_act_data=all_act_data.get(act)
            old_act_data.update(this_act_data)

        #print "dbc3:", len(dbconnection.queries)
        return all_act_data

    #---------------------------------------------------------------------------
    def sort_possibles(self, con, all_act_data,level1pairs,prefix_base):
        """Meant to make Otto run faster by moving all db queries.
        Takes in dict w/k of activity, v list of criteria, including possible occurrences
        now need to sort by interest match, conflicts, among self and each other, without too maky db hits"""
        #print "starting sort possibilities"
        #print "dbc1:", len(dbconnection.queries)
        from swingtime.forms import L1Check
        #date_range=con.get_date_range()
        figureheads=[]
        participants=[]
        fpks=[]
        ppks=[]
        busy={}
        busy_coaching={}

        #print "dbc2:", len(dbconnection.queries)
        for attr_dict in all_act_data.values():
            these_f=attr_dict.get('figureheads')
            for f in these_f:
                if f not in figureheads:
                    figureheads.append(f)
            these_p=attr_dict.get('participants')
            for p in these_p:
                if p not in participants:
                    participants.append(p)

        for f in figureheads:
            if f.pk not in fpks:
                fpks.append(f.pk)
            if f not in busy:
                busy[f]=[]

        for p in participants:
            if p.pk not in ppks:
                ppks.append(p.pk)
            if p not in busy:
                busy[p]=[]

        #print "dbc3 models:", len(dbconnection.queries)
        scheduled_os=Occurrence.objects.filter(start_time__gte=con.start, end_time__lte=con.end).exclude(training=None,challenge=None).prefetch_related('training').prefetch_related('training__coach__user__registrant_set').prefetch_related('challenge').prefetch_related('challenge__roster1__participants').prefetch_related('challenge__roster2__participants')
        #print "dbc4:", len(dbconnection.queries)
        #1 chal=8 db hits above. too many? 2 trains=7 hits maybe doesn't matter how many acts?
        for o in scheduled_os:
            if o.challenge:
                for roster in [o.challenge.roster1, o.challenge.roster2]:
                    for r in roster.participants.all():
                        if r in busy:
                            r_busy=busy.get(r)
                            r_busy.append(o)
                            busy[r]=list(r_busy)
                        else:
                            busy[r]=[o]

            elif o.training:
                for c in o.training.coach.all():
                    for r in c.user.registrant_set.filter(con=con):
                        if r in busy_coaching:
                            r_busy=busy_coaching.get(r)
                            r_busy.append(o)
                            busy_coaching[r]=list(r_busy)
                        else:
                            busy_coaching[r]=[o]
        #first make a separate coaching dict for conflict checks later, then conbine with the regular busy dict
        for k,v in busy_coaching.iteritems():
            if k in busy:
                r_busy=busy.get(k)
                for o in v:
                    if o not in r_busy:
                        r_busy.append(o)
                busy[k]=list(r_busy) #i dobn't actually think this is necessary
            else:
                busy[k]=v

        #print "dbc5:", len(dbconnection.queries)
        related_blackouts=Blackout.objects.filter(registrant__in=figureheads).prefetch_related('registrant')
        #print "dbc6:", len(dbconnection.queries)

        for b in related_blackouts:
            r_busy=busy.get(b.registrant)
            tempo=b.make_temp_o()
            r_busy.append(tempo)
            busy[b.registrant]=list(r_busy)


        avail_score_dict={}
        for act,this_act_dict in all_act_data.iteritems():
            print "\nact: ",act
            act_p=this_act_dict.get('participants')
            act_f=this_act_dict.get('figureheads')

            level1=[]
            level15=[]
            level2=[]
            interestexact=this_act_dict.get("interestexact")
            interestremoved=this_act_dict.get("interestremoved")

            #check to see if any participants are coaches, and if so, remove coach session occurrences
            coach_participants=set(act_p).intersection(set(busy_coaching.keys()))
############################## uncomment here if you want to be less restrictive w/ coach schedules, only get rid of an occurrence if they're busy coaching at the time.
    ## solution 1
            # if len(coach_participants)>0:
            #     for l in [interestexact,interestremoved]:
            #         to_remove=[]
            #         for o in l:
            #             coach_intersect=o.busy_soft(list(coach_participants),busy_coaching)
            #             if coach_intersect:
            #                 to_remove.append(o)
            #                 ###this is the problem. after it removes one it skips the next in line
            #                 #l.remove(o)
            #                 #don't change the to_remove instead of l.remove, this was making a mystery by not testing all occurrences against coach schedules, was skipping
            #
            #         while len(to_remove)>0:
            #             for o in to_remove:
            #                 l.remove(o)
            #                 to_remove.remove(o)

## solution 2
            #this is actually a mcuh easier way to deal with that.
            #adding any coach as a figurehead makes sure they have no conflicts.
            #doesn't distinguish betwheen whether they're coaching at the time or not.
            #might cause a problem by being too restrictive, giving coaches too much preference.
            #fuck it, I'll leave it that way for now.
            for c in list(coach_participants):
                if c not in act_f:
                    act_f.append(c)
###############################################################
            for o in interestexact:
                figurehead_intersect=o.busy_soft(act_f,busy)
                participant_intersect=o.busy_soft(act_p,busy)

                if not figurehead_intersect:
                    if not participant_intersect:
                        level1.append(o)
                    else:
                        level2.append(o)

            for o in interestremoved:
                figurehead_intersect=o.busy_soft(act_f,busy)
                participant_intersect=o.busy_soft(act_p,busy)

                if not figurehead_intersect:
                    if not participant_intersect:
                        level15.append(o)
                    else:
                        level2.append(o)

            avail_score=( (len(level1)*100) + (len(level15)*10) + (len(level2)*1) )
            if avail_score in avail_score_dict:
                tmp=avail_score_dict.get(avail_score)
                tmp.append(act)
            else:
                avail_score_dict[avail_score]=[act]

            this_act_dict.update({"level1":level1,"level15":level15,"level2":level2,"avail_score":avail_score})

        #print "all_act_data test",all_act_data

        taken_os=[]
        #print "avail_score_dict",avail_score_dict
        ask=avail_score_dict.keys()
        ask.sort()#sorting less available to more available
        #print "ask",ask
        for score in ask:
            #print "score",score
            act_list=avail_score_dict.get(score)
            #print "act_list",act_list
            for act in act_list:
                #print "\nact: ",act
                oselected=False
                this_act_dict=all_act_data.get(act)
                l1=this_act_dict.get('level1')
                #print "len l1: ",len(l1)
                l15=this_act_dict.get('level15')
                #print "len l15: ",len(l15)
                l2=this_act_dict.get('level2')
                #print "len l2: ",len(l2)
                figs=this_act_dict.get('figureheads')
                parts=this_act_dict.get('participants')
                #print "len figs",len(figs)
                #print "len partss",len(parts)


                if len(l1)>0:
                    while len(l1)>0 and not oselected:
                        o=choice(l1)
                        #print"o: ",o.start_time,o.end_time,o.interest#to see if the break is working
                        if o not in taken_os:
                            figurehead_intersect=o.busy_soft(figs,busy)
                            if not figurehead_intersect:
                                participant_intersect=o.busy_soft(parts,busy)
                                if not participant_intersect:
                                    #print"not in taken"
                                    prefix=prefix_base+"-%s-occurr-%s"%(str(act.pk),str(o.pk))
                                    #print"prefix: ",prefix
                                    level1pairs[(act,o,"Perfect Match")]=L1Check(prefix=prefix)
                                    #print "selecting ", o.start_time, o.end_time
                                    taken_os.append(o)
                                    l1.remove(o)
                                    oselected=True

                                    for l in [figs,parts]:
                                        for r in l:
                                            r_busy=busy.get(r)
                                            r_busy.append(o)
                                    break
                                else:#if participant intersect
                                    l1.remove(o)
                                    if o not in l2:
                                        l2.append(o)
                            else:
                                #print "o taken, keep going"
                                #print "len l1: ",len(l1)
                                l1.remove(o)
                        else:
                            #print "o taken, keep going"
                            #print "len l1: ",len(l1)
                            l1.remove(o)

                if len(l15)>0 and not oselected:
                    while len(l15)>0 and not oselected:
                        o=choice(l15)
                        #print"o: ",o.start_time,o.end_time,o.interest#to see if the break is working
                        if o not in taken_os:
                            figurehead_intersect=o.busy_soft(figs,busy)
                            if not figurehead_intersect:
                                participant_intersect=o.busy_soft(parts,busy)
                                if not participant_intersect:
                                    #print"not in taken"
                                    prefix=prefix_base+"-%s-occurr-%s"%(str(act.pk),str(o.pk))
                                    #print"prefix: ",prefix
                                    level1pairs[(act,o,"+/- Interest but no Conflicts")]=L1Check(prefix=prefix)
                                    taken_os.append(o)
                                    l15.remove(o)
                                    oselected=True
                                    #print "selecting ", o.start_time, o.end_time
                                    for l in [figs,parts]:
                                        for r in l:
                                            r_busy=busy.get(r)
                                            r_busy.append(o)
                                    break
                                else:#if participant intersect
                                    l1.remove(o)
                                    if o not in l2:
                                        l2.append(o)
                            else:
                                #print "o taken, keep going"
                                #print "len l15: ",len(l15)
                                l15.remove(o)
                                if not figurehead_intersect:
                                    if o not in l2:
                                        l2.append(o)
                        else:
                            #print "o taken, keep going"
                            #print "len l15: ",len(l15)
                            l15.remove(o)

                elif len(l2)>0 and not oselected:
                    #print "going for l2, len ",len(l2)
                    while len(l2)>0 and not oselected:
                        o=choice(l2)
                        #print"o: ",o.start_time,o.end_time,o.interest#to see if the break is working
                        if o not in taken_os:
                            figurehead_intersect=o.busy_soft(figs,busy)
                            if not figurehead_intersect:
                                #print"not in taken"
                                prefix=prefix_base+"-%s-occurr-%s"%(str(act.pk),str(o.pk))
                                #print"prefix: ",prefix
                                level1pairs[(act,o,"+/- Interest and Player Conflicts")]=L1Check(prefix=prefix)
                                taken_os.append(o)
                                #print "selecting ", o.start_time, o.end_time
                                l2.remove(o)
                                oselected=True

                                for l in [figs,parts]:
                                    for r in l:
                                        r_busy=busy.get(r)
                                        r_busy.append(o)
                                break
                            else:
                                #print "o taken, keep going"
                                #print "len l2: ",len(l2)
                                l2.remove(o)
                        else:
                            #print "o taken, keep going"
                            #print "len l2: ",len(l2)
                            l2.remove(o)

        return level1pairs




#===============================================================================
@python_2_unicode_compatible
class Occurrence(models.Model):
    '''
    Represents the start end time for a specific occurrence of a master ``Event``
    object.
    '''
    start_time = models.DateTimeField(_('start time'))
    end_time = models.DateTimeField(_('end time'))
    #Why did i say null=true for locaiton? prob so occurrance doesn't get deleted if locaiton does
    location = models.ForeignKey(Location, null=True, blank=True, on_delete=models.SET_NULL)
    interest=models.IntegerField(null=True, blank=True,choices=INTEREST_RATING, default=3)

    training=models.ForeignKey(Training,null=True,blank=True,on_delete=models.SET_NULL)
    challenge=models.ForeignKey(Challenge,null=True,blank=True,on_delete=models.SET_NULL)

    objects = OccurrenceManager()

    #===========================================================================
    class Meta:
        verbose_name = _('occurrence')
        verbose_name_plural = _('occurrences')
        ordering = ('start_time', 'end_time')

    #---------------------------------------------------------------------------
    def __str__(self):
        #return '{}: {}'.format(self.title, self.start_time.isoformat())
        return '{}: {}'.format(self.name, self.start_time.isoformat())

    #---------------------------------------------------------------------------
    @models.permalink
    def get_absolute_url(self):
        return ('swingtime-occurrence', [str(self.id)])

    #---------------------------------------------------------------------------
    def __lt__(self, other):
        return self.start_time < other.start_time


    #---------------------------------------------------------------------------
    def get_activity(self):
        '''
        Dahmer custom. Returns training or activity related to Event.
        Currently written to ease the transition to not having Event as a model
        '''
        activity=None

        if self.training or self.challenge:
            if self.training:
                activity=self.training
            elif self.challenge:
                activity=self.challenge

        return activity
    #---------------------------------------------------------------------------
#######eventually I want this to replace event model, just being cautious######################
    @property
    def activity(self):
        return self.get_activity()


    #---------------------------------------------------------------------------
########eventually I want this to replace title######################
    @property
    def name(self):
        activity=self.get_activity()

        if activity and activity.name:
            return activity.name
        else:
            if self.interest:
                temp_name="Empty (Interest: "+str(self.interest)+")"
            else:
                temp_name="Empty"
            return temp_name
    #---------------------------------------------------------------------------

    #reference: http://stackoverflow.com/questions/7366363/adding-custom-django-model-validation

    def validate_unique(self, *args, **kwargs):
        super(Occurrence, self).validate_unique(*args, **kwargs)

        if self.start_time and self.end_time and self.location:

            qs = self.__class__._default_manager.filter(
                start_time__lt=self.end_time,
                end_time__gt=self.start_time,
                location=self.location)

            if not self._state.adding and self.pk is not None:
                qs = qs.exclude(pk=self.pk)

            if qs.exists():
                raise ValidationError({
                    NON_FIELD_ERRORS: ["You can't have more than 1 occurrence in the same place at the same time",],})

        if self.training and self.challenge:#can't add until occurrence has training and challenge
            raise ValidationError({
                NON_FIELD_ERRORS: ["Occurrence cannot be BOTH a Challenge and a Training",],})



    #---------------------------------------------------------------------------
    def get_endtime(self):
        """Dahmer custom funct to get end time based on the train/chal it's linked to,
        and pad an extra 15 mins for 30 min chal or 30 mins for 90 min chal
        REQUIRES event and start time"""
        duration=1#default, may be overridden
        activity=self.get_activity()
        if activity and activity.duration:
            duration=float(activity.duration)

        dur_delta=int(duration*60)
        end_time=self.start_time+datetime.timedelta(minutes=dur_delta)
        return end_time
    #---------------------------------------------------------------------------
    def get_initial_interest(self,start_time):
        """Figures out intiial interest level based on when event is occurring.
        Start time is fed so can be calculated for initial before any other data is populated."""
        pass

    #---------------------------------------------------------------------------
    def figurehead_conflict(self):
        """Dahmer custom function. Checks to see if any Event figureheads (coach, captain) are participating in other occurrances at same time.
        If so, returns dict w/  activity k, skater list v, but has no teeth, can be overridden, is just an FYI warning"""
        activity=self.get_activity()#0db hits
        figureheads=[]
        conflict_dict={}
        if activity:
            figureheads=activity.get_figurehead_registrants()#figureheads is for getting blackouts, but where to put the logic?
        else:
            figureheads=[]

        #0 db hits
        #concurrent=Occurrence.objects.filter(start_time__lt=self.end_time,end_time__gt=self.start_time).exclude(pk=self.pk).select_related('challenge').select_related('training')
        #adding a half hour padding:
        concurrent=list(Occurrence.objects.filter(start_time__lt=(self.end_time + datetime.timedelta(minutes=30)),end_time__gt=(self.start_time - datetime.timedelta(minutes=30))).exclude(pk=self.pk).select_related('challenge').select_related('training'))

        for o in concurrent:
            event_activity=o.get_activity()
            if event_activity: #could be an empty timeslot
                event_part=event_activity.participating_in()
                conflict=set(figureheads).intersection(event_part)
                if len( conflict ) > 0:
                    conflict_dict[event_activity]=list(conflict)

        if len(conflict_dict)>0:
            return conflict_dict
        else:
            return None
    #---------------------------------------------------------------------------
    def os_soft_intersect(self,o2):
        #print "starting os_soft_intersect: ", self.start_time, self.end_time
        #print "against ",o2.start_time, o2.end_time
        if (o2.start_time<(self.end_time + datetime.timedelta(minutes=30))) and (o2.end_time>(self.start_time - datetime.timedelta(minutes=30) ) ):
            #print "true"
            return True
        else:
            #print "false"
            return False
    #---------------------------------------------------------------------------
    def os_hard_intersect(self,o2):
        #print "starting os_hard_intersect: ", self.start_time, self.end_time
        #print "against ",o2.start_time, o2.end_time
        if (o2.start_time<self.end_time) and (o2.end_time>self.start_time):
            #print "true"
            return True
        else:
            #print "false"
            return False
    #-------------------------------------------------------------------------------
    def busy_soft(self,participant_list,busy_dict):
        """takes in list of relevant registrant, dict w/ registrant as key, list of occurrences reg is in as v,
        checks to see if reg is busy/ w/ soft intersection"""
        #print "\n\nstarting  busy_soft for ",self.challenge,self.training,self.start_time,self.end_time
        intersection=False
        i=0
        #print "len(participant_list)",len(participant_list)
        while not intersection and i<len(participant_list):
            for f in participant_list:
                #print "starting i is ",i
                #print f
                busy_list=busy_dict.get(f)
                #print "busy_list",busy_list
                for o2 in busy_list:
                    #print "tessting ",o2.start_time,o2.end_time
                    if self.os_soft_intersect(o2):
                        intersection=True
                        #print "about to break a"
                        break
                i+=1
                if intersection:
                    #print "about to break b"
                    break
            if intersection:
                #print "about to break c"
                break
            #print "ending i is ",i
        #print "busy_soft intersection",intersection
        return intersection

    #-------------------------------------------------------------------------------
    def participant_conflict(self):
        """Dahmer custom function. Checks to see if any Event participants are participating in other occurrances at same time.
        If so, returns dict w/  activity k, skater list v, but has no teeth, can be overridden, is just an FYI warning"""
        activity=self.get_activity()
        figureheads=[]
        conflict_dict={}

        if activity:
            occur_part = activity.participating_in()#cut down form 7 to 4 db hits
        else:
            occur_part=[]
        #concurrent=Occurrence.objects.filter(start_time__lt=self.end_time,end_time__gt=self.start_time).exclude(pk=self.pk)
        #adding a half hour padding:
        concurrent=Occurrence.objects.filter(start_time__lt=(self.end_time + datetime.timedelta(minutes=30)),end_time__gt=(self.start_time - datetime.timedelta(minutes=30))).exclude(pk=self.pk)

        for o in concurrent:
            event_activity=o.get_activity()
            if event_activity: #could be an empty timeslot
                event_part=event_activity.participating_in()
                inter=set(occur_part).intersection(event_part)
                if len( inter ) > 0:
                    conflict_dict[event_activity]=list(inter)
                    #conflict_dict[event_activity]=event_part#whoops, this adds all people, not just the intersection!
        if len(conflict_dict)>0:
            #print "conflict_dict",conflict_dict
            return conflict_dict
        else:
            return None

    #-------------------------------------------------------------------------------
    def blackout_conflict(self):
        """Dahmer custom function. Checks to see if any Event figureheads (coach, captain have blackouts during Occu time.
        If so, returns dict w/  blackout k, skater list v, but has no teeth, can be overridden, is just an FYI warning"""
        #print("running o blackout_conflict")
        activity=self.get_activity()
        figureheads=[]
        conflict_dict={}
        daypart=[]
        #print("activity ",activity)
        if activity:
            figureheads=activity.get_figurehead_registrants()#figureheads is for getting blackouts, but where to put the logic?
        else:
            figureheads=[]

        odate=self.start_time.date()
        if self.start_time.time() >= parse_time('12:00:00'):
            daypart.append("PM")
        else:
            daypart.append("AM")
        if (self.end_time.time() >= parse_time('12:00:00')) and ("PM" not in daypart):
            daypart.append("PM")
        #print("daypart ",daypart)
        for f in figureheads:
            #print("Figurehead ",f)
            potential_bouts=Blackout.objects.filter(registrant=f, date=odate, ampm__in=daypart)
            if len(list(potential_bouts))>0:
                conflict_dict[f]=list(potential_bouts)
        #print("conflict_dict ",conflict_dict)

        if len(conflict_dict)>0:
            return conflict_dict
        else:
            return None
#-------------------------------------------------------------------------------
    def get_add_url(self):
        """Creates string of URL to call add-event page.
        Like a DIY get abdolute url"""

        dstr_str=self.start_time.isoformat()

        url_str='/events/add/?dtstart=%s&location=%s'%(dstr_str,str(self.location.pk))

        if self.training:
            url_str+="&training=%s"%(str(self.training.pk))
        if self.challenge:
            url_str+="&challenge=%s"%(str(self.challenge.pk))

        return url_str

#-------------------------------------------------------------------------------
    def can_add_sk8ers(self):
        '''returns list of Users that can edit Roster
        this is for adding/removing roster participants,
        they are only true in activity.editable_by. This is Bosses and Volunteers.'''
        allowed_editors=list(User.objects.filter(groups__name__in=['Volunteer',BIG_BOSS_GROUP_NAME,LOWER_BOSS_GROUP_NAME]))
        return allowed_editors

#-------------------------------------------------------------------------------
    def excel_backup(self):
        """Writes/returns xlsx file of relevant data, to be gathered ahead of the con and be used as a Plan B,
        or for people who don't want to use the site, I guess (I think redundant, Ivanna insists)"""
        wb=None

        if self.training or self.challenge:

            timestr=self.start_time.strftime('%H %M %p ')
            xlfilename=timestr+(ascii_only_no_punct(self.name))+".xlsx"

            wb = openpyxl.Workbook()
            sheet = wb.active

            if self.training:

                regros, rcreated=TrainingRoster.objects.get_or_create(registered=self)

                if regros.intl:
                    splitxlfilename=xlfilename.split(".")
                    newname=splitxlfilename[0]+" INTL"
                    xlfilename=''.join([newname,splitxlfilename[1]])

                sheet["E3"].value = "Printed: %s"%(timezone.now().strftime('%m %d %Y'))
                sheet["A1"].value = self.training.name
                sheet["A2"].value = self.training.display_coach_names()
                sheet["A3"].value = self.start_time.strftime('%H %M %p, %m-%d-%Y')
                sheet["A4"].value = self.location.name

                if regros.intl:
                    sheet["B6"].value = "INTL ONLY Registration Roster"
                else:
                    sheet["B6"].value = "Registration Roster"
                sheet["A7"].value = "Skill:"
                sheet["B7"].value = self.training.skill_display()
                sheet["A8"].value = "Pass:"
                sheet["B8"].value = self.training.passes_str()
                sheet["B10"].value = "Name"

                if regros.intl:
                    sheet["F6"].value = "Auditing Roster (non-INTL skaters go here, maybe coach will let them skate.)"
                else:
                    sheet["F6"].value = "Auditing Roster"

                sheet["E7"].value = "Skill:"
                sheet["F7"].value = "ABCD"
                sheet["E8"].value = "Pass:"
                sheet["F8"].value = "MVP, Skater, Offskate" #right? confirm.
                sheet["F10"].value = "Name"

                starti=11
                rno=int(1)
                #regros, rcreated=TrainingRoster.objects.get_or_create(registered=acto) #moved above so i can find out if INTL
                rmaxcap=regros.get_maxcap()
                for r in range(1,(rmaxcap+1)):
                    sheet["A"+str(starti)].value = str(rno)+"."
                    rno+=1
                    starti+=1

                starti=11
                rno=int(1)
                audros,acreated=TrainingRoster.objects.get_or_create(auditing=self)
                amaxcap=audros.get_maxcap()
                for r in range(1,(amaxcap+1)):
                    sheet["E"+str(starti)].value = str(rno)+"."
                    rno+=1
                    starti+=1

            elif self.challenge:

                sheet["A1"].value = self.challenge.data_title

                sheet["A2"].value = self.location.abbrv
                sheet["B2"].value = self.start_time.strftime('%H %M %p, %m-%d-%Y')

                sheet["D2"].value = "Printed:"
                sheet["E2"].value = timezone.now().strftime('%m %d %Y')

                if self.challenge.communication:
                    sheet["A4"].value = "ATTN:"
                    sheet["A4"].fill = PatternFill(start_color='FFFF0000',end_color="FFFFFF00",fill_type='solid')
                    sheet.merge_cells('B4:G4')
                    sheet["B4"].value = self.challenge.communication
                    sheet.row_dimensions[4].height=(12 * len( self.challenge.communication.splitlines() ) )

                sheet["A6"].value = "TEAM"
                sheet["B6"].value = self.challenge.roster1.name
                sheet["A7"].value = "COLOR"
                sheet["B7"].value = self.challenge.roster1.color
                sheet["A8"].value = "CAPTAIN"
                sheet["B8"].value = self.challenge.roster1.captain.name
                if self.challenge.roster1.captain.email:
                    sheet["C8"].value = self.challenge.roster1.captain.email
                sheet["A9"].value = "SKILL"
                sheet["B9"].value = self.challenge.roster1.skill_display()
                sheet["A10"].value = "GENDER"
                sheet["B10"].value = self.challenge.roster1.gender_text()

                sheet["E6"].value = "TEAM"
                sheet["E7"].value = "COLOR"
                sheet["F6"].value = self.challenge.roster2.name
                sheet["F7"].value = self.challenge.roster2.color
                sheet["E8"].value = "CAPTAIN"
                sheet["F8"].value = self.challenge.roster2.captain.name
                if self.challenge.roster2.captain.email:
                    sheet["G8"].value = self.challenge.roster2.captain.email

                sheet["E9"].value = "SKILL"
                sheet["F9"].value = self.challenge.roster2.skill_display()
                sheet["E10"].value = "GENDER"
                sheet["F10"].value = self.challenge.roster2.gender_text()

                sheet["A12"].value = "# of players"
                sheet["E12"].value = "# of players"
                sheet["B12"].value = "Skater #"
                sheet["F12"].value = "Skater #"
                sheet["C12"].value = "Skater Name"
                sheet["G12"].value = "Skater Name"

                starti=13
                rno=int(1)
                r1=list(self.challenge.roster1.participants.all())
                r1.sort(key=lambda x: x.sk8number)
                for r in r1:
                    if r==self.challenge.roster1.captain:
                        if r.sk8name:
                            name=r.sk8name+" (Captain)"
                        else:
                            name="(Captain)"
                    else:
                        name=r.sk8name
                    sheet["A"+str(starti)].value = str(rno)+"."
                    sheet["B"+str(starti)].value = r.sk8number
                    sheet["C"+str(starti)].value = name
                    rno+=1
                    starti+=1

                starti=13
                rno=int(1)
                r2=list(self.challenge.roster2.participants.all())
                r2.sort(key=lambda x: x.sk8number)
                for r in r2:
                    if r==self.challenge.roster2.captain:
                        if r.sk8name:
                            name=r.sk8name+" (Captain)"
                        else:
                            name="(Captain)"
                    else:
                        name=r.sk8name
                    sheet["E"+str(starti)].value = str(rno)+"."
                    sheet["F"+str(starti)].value = r.sk8number
                    sheet["G"+str(starti)].value = name
                    rno+=1
                    starti+=1

        return wb,xlfilename

#-------------------------------------------------------------------------------
class TrainingRoster(models.Model):
    """Used for Registration and Auditing roster for training Occurrences.
    Can be made in the Admin if made INTL, or made using get_or_create in register_training view.
    Otherwise, not necessary yet."""

    #gender=models.CharField(max_length=30, choices=GENDER, default=GENDER[0][0])
    skill=models.CharField(max_length=30, null=True,blank=True,choices=SKILL_LEVEL_TNG)#make this so that it cn be same as training or not, or maube just get rid of it
    intl=models.NullBooleanField(default=False)
    participants=models.ManyToManyField(Registrant, blank=True)
    cap=models.IntegerField(null=True, blank=True)

    #Reminder: can have 1 of these but not both.
      #why did i decide to do it this way, instead of occurrence linking to TrainingRoster?
      #maybe so i don't bog down every non-training occurrence w/2 empty fields.
    registered=models.OneToOneField("Occurrence", null=True,blank=True,related_name="registered")
    auditing=models.OneToOneField("Occurrence", null=True,blank=True,related_name="auditing")

    def __unicode__(self):
        return self.name

    class Meta:
        ordering=('registered__start_time','auditing__start_time','registered__training__name','auditing__training__name')

    @property
    def name(self):
        basename=""
        if self.registered:
            if self.intl:
                basename+="INTL "
            basename+=("%s %s (REGISTERED)"%(self.registered.name,self.registered.start_time.strftime("%a %B %d %I:%-M %p")))
        elif self.auditing:
            basename+=("%s %s (AUDITING)"%(self.auditing.name, self.auditing.start_time.strftime("%a %B %d %I:%-M %p")))
        else:
            basename+="Training Roster sans Training"

        return basename
    #---------------------------------------------------------------------------
    def validate_unique(self, *args, **kwargs):
        super(TrainingRoster, self).validate_unique(*args, **kwargs)
        if self.registered and self.auditing:
            raise ValidationError({
                NON_FIELD_ERRORS: ["Roster cannot be both Registered and Auditing",],})

        if not self.registered and not self.auditing:
            raise ValidationError({
                NON_FIELD_ERRORS: ["Please choose a Training Occurrence",],})

    def intl_icon(self):
        if self.intl:
            #return "glyphicon icon-passportbig"
            #return "glyphicon icon-plane-outline"
            return "glyphicon icon-globe-alt"
        else:
            return "glyphicon icon-universal-access"

    def intl_text(self):
        if self.intl:
            return "International"
        else:
            return None

    def intl_tooltip_title(self):
        if self.intl:
            return "Registrant must qualify as 'International' in order to register. Any MVP can audit and non-INTL auditing skaters MIGHT be allowed to participate as if registered if space is available."
        else:
            return "No location restrictions for registration"

    def intls_allowed(self):
        if self.intl is True:
            allowed=[True]
        else:
            allowed=[True,False,None]
        return allowed

    def can_register_at(self):
        """Returns datetime registration for class starts
        both for displaying that time and checking if now is past that time"""

        can_reg=None
        regtimes=[]

        if self.registered or self.auditing:
            if self.registered:
                o=self.registered
            elif self.auditing:
                o=self.auditing
            con=o.training.con

            ostart=datetime.time(hour=o.start_time.hour,minute=o.start_time.minute)
            if ostart<=con.morning_class_cutoff:#if class starts early enough in the morning
                yday = datetime.timedelta(days=1)
                startday=o.start_time.date()-yday
                regtimes.append(datetime.datetime(year=startday.year, month=startday.month, day=startday.day, hour=con.dayb4signup_start.hour, minute=con.dayb4signup_start.minute))

            #otherwise this is time-hoursb4signup
            #it calculates both just in case timezone doesn't work and they do something like 48 hours before class time or something
            b4signup = datetime.timedelta(hours=float(con.hoursb4signup))
            regtime=o.start_time-b4signup
            regtimes.append(datetime.datetime(year=o.start_time.year, month=o.start_time.month, day=o.start_time.day, hour=regtime.hour, minute=regtime.minute))

            if len(regtimes)>0:
                can_reg=min(regtimes)

        return can_reg


    def can_register(self):
        """Returns true if registration window is open, False if not.
        Will be determined by 2(?) hour window before class starts"""
        can_reg=self.can_register_at()
        now=timezone.now()

        if can_reg and now>=can_reg:
            return True
        else:
            return False

    def get_maxcap(self):
        '''checks is roster has a cap. If not, supplies defaults listed at top of file.
        PRIORITY: self.cap for TrainingRoster, then if Training specifically has a regcap or audcap, then if not, general default cap listed in Scheduler.models.
        If this is the auditing roster of an INTL training, it allows the audit cap to be
        general training defaults-number of people registered. Tht is so coaches can have a larger audit roster in empty INTL classes.
        LOOPHOLE: people w/out an MVP pass can sign up to audit an INTL class and then be allowed in to participate.
        I think it'll take people a long time to figure tha tout, if they ever do.'''
        intl=False

        if self.registered and self.registered.training:
            #If this is a registration training roster

            #get regcap
            if self.cap:
                regcap=self.cap
            elif self.registered.training.regcap:
                regcap=self.registered.training.regcap
            else:
                regcap=DEFAULT_REG_CAP

            if self.intl:
                intl=True

                #get audcap. Only need if is intl
                audcap=DEFAULT_AUD_CAP #might get overwritten, jsut here so I don't need ot do 2 elses
                if self.registered.auditing and self.registered.auditing.cap:
                    audcap=self.registered.auditing.cap
                elif self.registered.training and self.registered.training.audcap:
                    audcap=self.registered.training.audcap
            else:#if not intl
                maxcap=regcap

        elif self.auditing and self.auditing.training:
            #if this is the auditing trainingroster
            #need to get audcap even if is INTL and gets overridden, bc is used in equation
            if self.cap:
                audcap=self.cap
            elif self.auditing.training.audcap:
                audcap=self.auditing.training.audcap
            else:
                audcap=DEFAULT_AUD_CAP

            if self.auditing.registered.intl:
                intl=True

                regcap=DEFAULT_REG_CAP#might get overwritten, jsut here so I don't need ot do 2 elses
                if self.auditing.registered and self.auditing.registered.cap:
                    regcap=self.auditing.registered.cap
                elif self.auditing.training and self.auditing.training.regcap:
                    regcap=self.auditingtraining.regcap
            else:
                maxcap=audcap

        if intl:
            if self.registered and self.registered.training:#If this is a registration training roster
            #what to do if is registration
                if self.registered.auditing:
                    audsk8=self.registered.auditing.participants.count()
                else:
                    audsk8=0

                intlborrow=audsk8-audcap #how mant people were borrowed from unfulfilled regcap?
                if intlborrow>0:
                    maxcap=regcap-intlborrow #how many people are allowed minus people borroed from reg roster for audit
                else:
                    maxcap=regcap

            elif self.auditing and self.auditing.training:#if this is the auditing trainingroster
                if self.auditing.registered:
                    regsk8=self.auditing.registered.participants.count()
                else:
                    regsk8=0

                maxcap=((regcap-regsk8)+audcap)

        return maxcap


    def spacea(self):
        '''gets maxcap (see above), checks is participants are fewer'''
        #print "starting spacea for ",self
        maxcap=self.get_maxcap()
        #print "maxcap ",maxcap
        spacea=maxcap-self.participants.count()
        #print "spacea ",spacea

        if spacea>0:
            return spacea
        else:
            return False

    def editable_by(self):
        '''returns list of Users that can edit Roster
        this is for adding/removing roster participants, so coaches actually don't have this permission,
        they are only true in activity.editable_by. This is Bosses and Volunteers.'''
        allowed_editors=list(User.objects.filter(groups__name__in=['Volunteer',BIG_BOSS_GROUP_NAME,LOWER_BOSS_GROUP_NAME]))
        return allowed_editors

    def skills_allowed(self):
        if self.skill:
            allowed=list(self.skill)
            if "O" in allowed:
                allowed.remove("O")
        else:
            allowed=["A","B","C","D"]
        return allowed

    def skill_display(self):
        """This makes it so I don't see A0, just A, or AB, or something more understandable"""
        prettify=''.join(self.skills_allowed())
        return prettify

    def skill_tooltip_title(self):
        if self.skill:
            allowed=self.skills_allowed()
            allowed.sort(reverse=True)
            skill_dict=dict(SKILL_LEVEL)
            str_base="Registrant must identify skill as"
            str_end= " in Profile in order to register"
            str_mid=""
            for item in allowed:
                if item:
                    displayable=skill_dict.get(item)
                    if item==allowed[-1]:
                        item_str=" or "+displayable
                    else:
                        item_str=" "+displayable+","

                str_mid+=item_str
            return str_base+str_mid+str_end
        else:
            return "No skill restrictions for registration"

    def skill_icon(self):
        if not self.skill:
            return "glyphicon icon-universal-access"

    # def save(self, *args, **kwargs):
    #     if self.registered and self.registered.training:
    #         if self.registered.training.regcap:# I think this might conflict with get_maxcap
    #             self.cap=self.registered.training.regcap#but maybe it almost never runs, since you'd have to add a regcap/audcap
    #     elif self.auditing and self.auditing.training:
    #         if self.auditing.training.audcap:# I think this might conflict with get_maxcap
    #             self.cap=self.auditing.training.audcap
    #     super(TrainingRoster, self).save()
