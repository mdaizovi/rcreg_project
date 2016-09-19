import collections
import datetime
import os
import openpyxl
from openpyxl import Workbook

from django import forms
from django.contrib.auth.models import Group, User
from django.core.exceptions import (
    ObjectDoesNotExist, ValidationError, NON_FIELD_ERRORS)
from django.db.models import Q
from django.utils.translation import ugettext_lazy as _

from con_event.models import Country, State, Con, Registrant
from rcreg_project.settings import BASE_DIR
from rcreg_project.extras import remove_punct, ascii_only, ascii_only_no_punct

"""Logic for uploading an XLSX sheet of registrants from the view 'upload_reg'.
XLSX must look exactly like it did in 2016."""

data_columns = [
    'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O',
    'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC',
    'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN']
base_header_file = os.path.join(BASE_DIR, 'con_event/BPT2016Header.xlsx')


class BPTUploadForm(forms.Form):
    """Used in 'upload_reg' to select XLSX file to upload."""

    def __init__(self, *args, **kwargs):
        super(BPTUploadForm, self).__init__(*args, **kwargs)
        CONS = []
        for con in Con.objects.all():
            CONS.append((con.pk, con))

        self.fields["con"] = (forms.CharField(widget=forms.Select(
                                choices=CONS),
                                label='Select Con',
                                initial=Con.objects.most_upcoming().pk)
                                )
        self.fields["xlfile"] = (forms.FileField(
                                label='Select File to Upload', required=True))

        for field in iter(self.fields):
            self.fields[field].widget.attrs.update({
                'class': 'form-control',
                })

    def get_header(self, header_file):
        """Gets header input xlsx file,
        returns as OrderedDict. {k: v} is {header: data} ."""

        wb = openpyxl.load_workbook(header_file)
        sheet = wb.get_active_sheet()
        header = collections.OrderedDict()

        for  c in data_columns:  # data_columns defined at top of file
            location = c + "1"
            data = str(sheet[location].value).strip().capitalize()
            header[c] = data

        return header


    def my_valid(self):
        """Don't touch this, it has to be this way in order to return
        specified custom error messages.
        Otherwise it loses uploaded file and thinks that's the problem.
        """

        IMPORT_FILE_TYPES = [".xls", ".xlsx", "application/vnd.ms-excel",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ]

        if self.is_valid():
            data = self.cleaned_data
            xlfile = data['xlfile']

            try:
                if xlfile:
                     if (xlfile.content_type and
                            xlfile.content_type not in IMPORT_FILE_TYPES):
                         self._errors["xlfile"] = (
                            self.error_class(['Please provide an Excel sheet'])
                            )

                     elif (self.get_header(base_header_file)
                            != self.get_header(xlfile)):
                         self._errors["xlfile"] = (
                            self.error_class(['The format of the uploaded  \
                            file, including the Header, must be identical  \
                            to 2016 BPT reports'])
                            )

                else:  # if not xlfile
                    self._errors["xlfile"] = (
                                        self.error_class(['Field Cannot be \
                                        blank.Please provide an Excel sheet.'])
                                        )
            except:
                self._errors["xlfile"] = (
                    self.error_class(["Unspecified error. \
                                    Please try another file."])
                                    )

            # if self.is_valid()
            return data

        else:  # if not self.is_valid()
            return False


    def make_excel_odict_list(self, xlfile):
        """Takes in XLSX file of BPT registrant data,
        turns each row into an OrderedDict,
        returns list of all OrderedDicts
        """

        wb = openpyxl.load_workbook(xlfile)
        sheet = wb.active
        all_data = []
        highest_row=sheet.get_highest_row()

        for row in range(2, sheet.get_highest_row() + 1):
            data_holding_dict = collections.OrderedDict()
            for  c in data_columns:
                location = c + str(row)
                data = sheet[location].value
                data_holding_dict[c] = data
            all_data.append(data_holding_dict)

        return all_data


    def find_incompletes(self, all_data):
        """Takes in OrderedDict list, presumably from  make_excel_odict_list,
        looks to see if any records are missing required data.
        """

        no_sk8_or_real_name = []
        complete_entries = []
        email_list = []
        last_email = None
        bad_emails = []

        def get_email(od):
            email = od.get("AB")
            if not email or len(email) < 5:  # 5 is arbitrary
                email = od.get("Q")
            return email

        # First get the emails on thir own, always same field.
        for od in all_data:
            email = get_email(od)
            if email:
                email_list.append(email)

        for od in all_data:
            sk8name = od.get("AC")
            first_name = od.get("AA")
            last_name = od.get("Z")
            email = get_email(od)

            if email:
                emailcount = email_list.count(email)

                if emailcount > 1:
                    bad_emails.append(od)
                elif email and sk8name and first_name and last_name:
                    complete_entries.append(od)
                else:
                    no_sk8_or_real_name.append(od)
            else:  # slight misnomer-- no email gets you here.
                no_sk8_or_real_name.append(od)

        return no_sk8_or_real_name, bad_emails, complete_entries


    def import_from_od(self, complete_entries, con):
        """Inputs list of complete_entries in form of OrderedDict,
        assumes already checked for email dupes, lack of real name, sk8name.
        This is coming from 2016 BPT format.
        """

        error_list = []
        success_list = []
        repeat_email_list = []

        for od in complete_entries:
            # Necessary information
            skill = od.get("AF")
            gender = od.get("AG")
            pass_type = od.get("V")
            email = od.get("AB")
            if not email:
                email = od.get("Q")
            first_name = od.get("AA")
            if not first_name:
                first_name = od.get("I")
            last_name = od.get("Z")
            if not last_name:
                last_name = od.get("H")

            # Unnecessary information, but desireable
            sk8name = od.get("AC")
            sk8number = str(od.get("AE"))
            country_name = od.get("O")
            state_name = od.get("M")
            BPT_Ticket_ID = od.get("A")

            # Yeah, sure, whatever. Maybe we'll do something with it someday.
            affiliation = od.get("AH")
            ins_carrier = od.get("AI")
            ins_number = od.get("AJ")
            age_group = od.get("AK")
            favorite_part = od.get("AM")
            volunteer = od.get("AL")

            for att in [
                    first_name, last_name, sk8name, sk8number, affiliation,
                    ins_carrier, ins_number, favorite_part, volunteer]:

                cleaned_att = ascii_only_no_punct(att)
                str_att = str(cleaned_att)

                if len(str_att) > 100:
                    att=str_att[:100]

            # Getting country and state determine INTL status automatically.
            if country_name:
                try:
                    country = Country.objects.get(name__iexact=country_name)
                except:
                    country = None
            else:
                country = None

            if state_name:
                try:
                    state = State.objects.get(slugname=state_name)
                except:
                    state = None
            else:
                state = None

            try:
                # This is the general try, any failure will send to error_list
                this_reg = None

                # Before adding registrant, make sure they don't already exist.
                try:
                    this_reg = (Registrant.objects.get(
                            con=con, email=email, first_name=first_name,
                            last_name=last_name, sk8name=sk8name)
                            )

                except ObjectDoesNotExist:
                    # First try:
                    # con and email match, and EITHER f/l name or ska8name
                    reg_q = (
                            Registrant.objects.filter(
                                con=con, email=email).filter(
                                Q(first_name__iexact=first_name,
                                    last_name__iexact=last_name) |
                                Q(sk8name__iexact=sk8name))
                            )

                    if reg_q.count() == 1:
                        this_reg = reg_q[0]

                    else:
                        # Allow for same person, different email
                        # IMPORTANT: real f, l, AND sk8 name, must all 3 match.
                        # Else, different people with same names can be merged.
                        # This happened in 2016.
                        reg_q = (
                                Registrant.objects.filter(
                                    con=con, first_name__iexact=first_name,
                                    last_name__iexact=last_name,
                                    sk8name__iexact=sk8name)
                                )

                        if reg_q.count() == 1:
                            this_reg = reg_q[0]

                        else:
                            try:
                                # If email already exists for con,
                                # but name doesn't match real or sk8 name
                                Registrant.objects.get(con=con, email=email)
                                repeat_email_list.append(od)

                            except ObjectDoesNotExist:
                                #Finally, I'll let you make a new registrant.
                                this_reg = (
                                        Registrant(
                                        con=con, email=email,
                                        first_name=first_name,
                                        last_name=last_name)
                                        )

                if this_reg:
                    # ie if no repeat email.
                    # Note, this will update into for an existing registrant.
                    attr_dict={
                        'sk8name':sk8name, 'sk8number':sk8number,
                        'skill':skill, 'gender':gender, 'pass_type':pass_type,
                        'first_name':first_name, 'last_name':last_name,
                        'country':country, 'state':state,
                        'BPT_Ticket_ID':BPT_Ticket_ID,
                        'affiliation':affiliation, 'ins_carrier':ins_carrier,
                        'ins_number':ins_number, 'age_group':age_group,
                        'favorite_part':favorite_part, 'volunteer':volunteer
                        }
                    for k, v in attr_dict.iteritems():
                        if v:
                            setattr(this_reg, k, v)

                    this_reg.save()
                    success_list.append(od)
            except:
                error_list.append(od)

        return success_list, error_list, repeat_email_list

    def make_registrants(self):
        """Makes resitrants from XLSX sheet upload,
        returns XLSX for download, notifying user (RC staff) of
        registrants made, and those that couldn't be made, for missing data
        """

        cdata = self.cleaned_data
        con = Con.objects.get(pk=int(cdata['con']))
        xlfile = cdata['xlfile']

        # Sort through the data
        header = self.get_header(base_header_file)
        all_data = self.make_excel_odict_list(xlfile)

        # Find the troublemakers
        no_sk8_or_real_name, bad_emails, complete_entries = (
                self.find_incompletes(all_data)
                )

        # Make the registrants you can, tell them about the ones you couldn't.
        reg_made, errors_list, email_dupe = (
                self.import_from_od(complete_entries, con)
                )

        bad_emails += email_dupe
        data_dict = collections.OrderedDict()
        data_dict["Registrants Made"] = reg_made
        data_dict["Errors"] = errors_list
        data_dict["Email Dupes"] = bad_emails
        data_dict["Incomplete Data"] = no_sk8_or_real_name

        # Output the data
        try:
            first = True
            wb = Workbook()
            for dk, dv in data_dict.iteritems():
                if first:
                    sheet = wb.get_active_sheet()
                    first = False
                else:
                    sheet = wb.create_sheet()
                sheet.title = dk

                r = 1
                for k, v in header.iteritems():
                    location = str(k) + str(r)
                    sheet[location].value = v
                for od in dv:
                    r += 1
                    for k, v in od.iteritems():
                        location = str(k) + str(r)
                        sheet[location].value = v
            return wb

        except:
            return None
